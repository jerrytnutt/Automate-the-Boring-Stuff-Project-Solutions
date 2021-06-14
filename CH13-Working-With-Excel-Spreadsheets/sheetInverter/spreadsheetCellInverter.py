# Write a program to invert the rows and columns of cells in spreadsheet.
# The value at row 5, column 3 will be at row 3, column 5 (and vice versa).

from openpyxl import load_workbook, Workbook

def invert_sheet():
  workbook = load_workbook('invertedSheetBefore.xlsx')
  sheet = workbook.active 
  
  max = sheet.max_row
  inventory_list = []
  # Create list with all cell values from before sheet
  for row_index in range(1,max):
    for cell in sheet[row_index]:
      inventory_list.append([cell.value, cell.row, cell.column])

  # Create new sheet to add inverted cells
  new_workbook = Workbook()
  new_sheet = new_workbook.active
  for items in inventory_list:
    if items[0] != None:
      value = items[0]
      column = items[1]
      row = items[2]
      new_sheet.cell(row=row, column=column).value = value
  new_workbook.save(filename="invertedSheetAfter.xlsx")

if __name__ == "__main__":
  invert_sheet()