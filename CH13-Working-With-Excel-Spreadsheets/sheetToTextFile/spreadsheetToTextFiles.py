# The program should open a spreadsheet and write the cells of column A into one text file, 
# The cells of column B into another text file, and so on.

from openpyxl import load_workbook
import os

def spread_to_text():
  workbook = load_workbook('spreadsheetToTextFiles.xlsx')
  sheet = workbook.active 
  inventory_list = []

  max_row = sheet.max_row
  max_column = sheet.max_column
  
  # Cycle the sheet row by row for each column
  for i in range(1,max_column + 1):
    file_name = "file" + str(i)+'.txt'
    for j in range(1,max_row + 1):
      with open(file_name, 'a') as myfile:
        # if the cells value is not empty append to the text file
        if sheet.cell(row=j, column=i).value != None:
          myfile.write(str(sheet.cell(row=j, column=i).value)+"\n")
     
if __name__ == "__main__":
  spread_to_text()