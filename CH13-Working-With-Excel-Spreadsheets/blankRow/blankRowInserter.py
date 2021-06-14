# Create a program that takes two integers and a filename. 
# With integer N and integer M, Start at row N and insert M amount of blank rows into the spreadsheet. 

from openpyxl import load_workbook, Workbook
import sys

def enter_blank_row(N,M):
  workbook = load_workbook('blankrowBefore.xlsx')
  sheet = workbook.active 
  N = int(N)
  M = int(M)
  
  # Store original rows into list
  row_list = []
  max = sheet.max_row
  for i in range(1,max + 1):
    row_list.append(list(sheet[i]))

  workbook = Workbook()
  new_sheet = workbook.active
  
  row = 1
  for items in row_list:
    column = 1
    for cell in items:
      new_sheet.cell(row=row, column=column).value = cell.value
      column += 1
    # Once row N is reached loop M times to enter blank rows
    if row == N:
      for i in range(M):
        row += 1
        new_sheet.cell(row=row, column=1).value = ''
    row += 1
  workbook.save('blankrowAfter.xlsx')

if __name__ == "__main__":
  if len(sys.argv) > 2:
    try:
      if int(sys.argv[1]) and int(sys.argv[2]) :
        enter_blank_row(sys.argv[1],sys.argv[2])
    except ValueError:
        print("Please enter two integers")