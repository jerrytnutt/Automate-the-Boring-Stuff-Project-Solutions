# Create a program that takes a number N from the command line
# And creates an N×N multiplication table in an Excel spreadsheet. 
from openpyxl import Workbook
from openpyxl.styles import Font
import sys

def create_table(n):
  workbook = Workbook()
  sheet = workbook.active
  n = int(n)

  # Add 2 to n on account of first labeled cell
  for i in range(1,n + 2):
    for j in range(1,n + 2):
      # loop through each cell and add the appropriate label or product
        if i == 1 and j != 1:
          sheet.cell(row=i, column=j).value = j - 1
          sheet.cell(row=i, column=j).font = Font(bold=True)
        elif i != 1 and j == 1:
          sheet.cell(row=i, column=j).value = i - 1 
          sheet.cell(row=i, column=j).font = Font(bold=True)
        else:
          sheet.cell(row=i, column=j).value = (i-1) * (j-1)
          
  sheet.cell(row=1, column=1).value = ''
  workbook.save(filename="multiplicationTable.xlsx")

if __name__ == "__main__":
  if len(sys.argv) > 1:
    try:
      if type(int(sys.argv[1])) == int:
        number = sys.argv[1]
        create_table(number)
    except ValueError:
      print("Please enter a number")
  else:
    print('Please enter a number')
  
 